import os
from openpyxl import load_workbook
from fpdf import FPDF
import matplotlib
import matplotlib.pyplot as plt
from os.path import join
filepath = "final_data.xlsx"
#create the workbook object and few inits
wb = load_workbook(filepath,data_only=True)
sheet = wb.active
max_row = sheet.max_row
max_column = sheet.max_column
cand = {}
#return the graph of the given data in the .png form to use in pdf
def bar_chart(id):
    bar_name1 = str(cand[i]["reg_no"])+"_1.png"
    bar_name2 = str(cand[i]["reg_no"]) + "_2.png"
    bar_name3 = str(cand[i]["reg_no"]) + "_3.png"
    data1 = {}
    data1[cand[i]["name"]] = cand[i]["t_marks"]
    data1["Average"] = cand[i]["avg"]
    data1["Median"] = cand[i]["median"]
    data1["Mode"] = cand[i]["mode"]
    data2 = {}
    data2[cand[i]["name"]] = cand[i]["attempt"]
    data2["World"] = cand[i]["avg_attempt"]
    data3 = {}
    data3[cand[i]["name"]] = cand[i]["accu"]
    data3["World"] = cand[i]["avg_accu"]
    candte = list(data1.keys())
    scores = list(data1.values())
    fig = plt.figure(figsize=(10, 5))
# creating the bar plot
    plt.bar(candte, scores)
    for j in range(len(candte)):
        plt.text(j, scores[j], scores[j], ha='center')
    plt.ylabel("Score")
    plt.title("Comparision of scores")
    plt.savefig(bar_name1, orientation='portrait', transparent=True, bbox_inches=None, pad_inches=0)
    plt.close()
    candte = list(data2.keys())
    scores = list(data2.values())
    fig = plt.figure(figsize=(10, 5))
    # creating the bar plot
    plt.bar(candte, scores)
    for j in range(len(candte)):
        plt.text(j, scores[j], scores[j], ha='center')
    plt.ylabel("Attempts(%)")
    plt.title("Comparision of Attempts(%)")
    plt.savefig(bar_name2, orientation='portrait', transparent=True, bbox_inches=None, pad_inches=0)
    plt.close()
    candte = list(data3.keys())
    scores = list(data3.values())
    fig = plt.figure(figsize=(10, 5))
    # creating the bar plot
    plt.bar(candte, scores)
    for j in range(len(candte)):
        plt.text(j, scores[j], scores[j], ha='center')
    plt.ylabel("Accuracy(%)")
    plt.title("Comparision of Accuracy(%)")
    plt.savefig(bar_name3, orientation='portrait', transparent=True, bbox_inches=None, pad_inches=0)
    plt.close()
#create the pdf using fpdf
def createPdf(id):
    pie_name = str(cand[i]["reg_no"]) + "_1.png"
    name = cand[i]["name"]
    ext = name + "("+str(cand[i]["reg_no"])+").pdf"
    desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
    report_cards = os.path.join(desktop,"report_cards")
    if not os.path.exists(report_cards):
        os.umask(0)
        os.makedirs(report_cards,mode=0o777)
    file_name = os.path.join(report_cards, ext)
    document = FPDF(unit='mm', format='A3')
    document.add_page(orientation="L")
    document.image('back.png', x=0, y=0, w=420, h=297)
    document.set_font('Arial','', 15)
    document.rect(5.0, 5.0, 410.0, 287.0)
    document.rect(8.0, 8.0, 404.0, 282.0)
    document.cell(20, 10, 'Round I - Enhanced Score Report : '+ str(cand[i]["name"]),ln=1)
    document.cell(20, 10, 'Reg Number : ' + str(cand[i]["reg_no"]), ln=1)
    document.cell(155)
    document.set_font('Arial', 'B', 40)
    document.cell(80, 50, 'PQR Championship Test',align="C",ln=1)
    document.cell(120)
    document.image("logo.png", w=150, h=70)
    document.set_font('Arial', 'B', 20)
    document.image(name+".png", 340, 20, w=50, h=50)
    document.cell(10)
    document.cell(0, 40, 'Round I performance of ' + str(cand[i]["name"]), ln=1)
    document.set_font('Arial', 'B', 15)
    document.cell(80)
    document.cell(60, 10, 'Grade',1,0)
    document.set_font('Arial', '', 15)
    document.cell(60, 10, str(cand[i]["grade"]),1,0)
    document.cell(20)
    document.set_font('Arial', 'B', 15)
    document.cell(60, 10, 'Registration No', 1, 0)
    document.set_font('Arial', '', 15)
    document.cell(60, 10, str(cand[i]["reg_no"]), 1, 1)
    document.cell(80)
    document.set_font('Arial', 'B', 15)
    document.cell(60, 10, 'School Name', 1, 0)
    document.set_font('Arial', '', 15)
    document.cell(60, 10, str(cand[i]["school_name"]), 1, 0)
    document.cell(20)
    document.set_font('Arial', 'B', 15)
    document.cell(60, 10, 'Gender', 1, 0)
    document.set_font('Arial', '', 15)
    document.cell(60, 10, str(cand[i]["gender"]), 1, 1)
    document.cell(80)
    document.set_font('Arial', 'B', 15)
    document.cell(60, 10, 'City Of Residence', 1, 0)
    document.set_font('Arial', '', 15)
    document.cell(60, 10, str(cand[i]["city"]), 1, 0)
    document.cell(20)
    document.set_font('Arial', 'B', 15)
    document.cell(60, 10, 'Date of Birth', 1, 0)
    document.set_font('Arial', '', 15)
    document.cell(60, 10, str(cand[i]["dob"]), 1, 1)
    document.cell(80)
    document.set_font('Arial', 'B', 15)
    document.cell(60, 10, 'Country Of Residence', 1, 0)
    document.set_font('Arial', '', 15)
    document.cell(60, 10, str(cand[i]["country"]), 1, 0)
    document.cell(20)
    document.set_font('Arial', 'B', 15)
    document.cell(60, 10, 'Date of Test', 1, 0)
    document.set_font('Arial', '', 15)
    document.cell(60, 10, str(cand[i]["dot"]), 1, 1)
    document.add_page(orientation="L")
    document.image('back.png', x=0, y=0, w=420, h=297)
    document.set_font('Arial', 'B', 15)
    document.rect(5.0, 5.0, 410.0, 287.0)
    document.rect(8.0, 8.0, 404.0, 282.0)
    document.cell(180)
    document.set_font('Arial', 'B', 20)
    document.cell(40, 20, 'Section 1', align="C", ln=1)
    document.cell(180)
    document.set_font('Arial', '', 15)
    document.cell(40, 20, "This section describes"+str(cand[i]["name"])+"'s performance v/s the Test in Grade "+str(cand[i]["grade"]), align="C", ln=1)
    document.ln(15)
    r_count = 0
    data = cand[i]["questions"]
    col_width = document.w / 8
    row_height = document.font_size*1.1
    for row in data:
        document.cell(15)
        if r_count == 0:
            document.set_text_color(255, 255, 255)
            for item in row:
                document.cell(col_width, row_height * 2, txt=str(item), border=1,fill=True)
            document.ln(row_height * 1)
        else:
            for item in row:
                document.set_text_color(0, 0, 0)
                document.cell(col_width, row_height * 1,
                              txt=str(item), border=1)
        document.ln(row_height * 1)
        r_count +=1
    document.cell(338)
    document.set_font('Arial', 'I', 20)
    document.cell(40, 20, 'Total Score: '+str(cand[i]["t_marks"]), align="C", ln=1)
#2nd sec
    document.add_page(orientation="L")
    document.image('back.png', x=0, y=0, w=420, h=297)
    document.set_font('Arial', 'B', 15)
    document.rect(5.0, 5.0, 410.0, 287.0)
    document.rect(8.0, 8.0, 404.0, 282.0)
    document.cell(180)
    document.set_font('Arial', 'B', 20)
    document.cell(40, 20, 'Section 2', align="C", ln=1)
    document.cell(180)
    document.set_font('Arial', '', 15)
    document.cell(40, 20, "This section describes "+str(cand[i]["name"])+"'s performance v/s the Rest of the World in Grade "+str(cand[i]["grade"]), align="C", ln=1)
    document.ln(10)
    document.set_font('Arial', '', 10)
    r_count = 0
    data = cand[i]["sec_2"]
    col_width = document.w / 20
    row_height = document.font_size * 1.1
    document.set_text_color(255, 255, 255)
    document.cell(40, 10, "Q.no",border=1, fill=True,ln=0,align="C")
    document.cell(40, 10, "Attempt Status", border=1, fill=True, ln=0,align="C")
    document.cell(40, 10,str(cand[i]["name"])+"'s" , border=1, fill=True, ln=0,align="C")
    document.cell(40, 10, "Correct", border=1, fill=True, ln=0,align="C")
    document.cell(40, 10, "Outcome", border=1, fill=True, ln=0,align="C")
    document.cell(40, 10, str(cand[i]["name"])+"'s", border=1, fill=True, ln=0,align="C")
    document.cell(40, 10, "% of students across the", border=1, fill=True, ln=0,align="C")
    document.cell(40, 10, "%of students (from those", border=1, fill=True, ln=0,align="C")
    document.cell(40, 10, "%of students(from those", border=1, fill=True, ln=0,align="C")
    document.cell(40, 10, "World Average", border=1, fill=True, ln=1,align="C")
    document.cell(40, 10, "", border=1, fill=True, ln=0,align="C")
    document.cell(40, 10, "", border=1, fill=True, ln=0,align="C")
    document.cell(40, 10, "Choice", border=1, fill=True, ln=0,align="C")
    document.cell(40, 10, "answer", border=1, fill=True, ln=0,align="C")
    document.cell(40, 10, "", border=1, fill=True, ln=0,align="C")
    document.cell(40, 10, "Score", border=1, fill=True, ln=0,align="C")
    document.cell(40, 10, "world who attempted ", border=1, fill=True, ln=0,align="C")
    document.cell(40, 10, "who attempted this ) who", border=1, fill=True, ln=0,align="C")
    document.cell(40, 10, "who attempted this) who", border=1, fill=True, ln=0,align="C")
    document.cell(40, 10, "in this question", border=1, fill=True, ln=1,align="C")
    document.cell(40, 10, "", border=1, fill=True, ln=0,align="C")
    document.cell(40, 10, "", border=1, fill=True, ln=0,align="C")
    document.cell(40, 10, "", border=1, fill=True, ln=0,align="C")
    document.cell(40, 10, "", border=1, fill=True, ln=0,align="C")
    document.cell(40, 10, "", border=1, fill=True, ln=0,align="C")
    document.cell(40, 10, "", border=1, fill=True, ln=0,align="C")
    document.cell(40, 10, "this question", border=1, fill=True, ln=0,align="C")
    document.cell(40, 10, "got it correct", border=1, fill=True, ln=0,align="C")
    document.cell(40, 10, "got it incorrect", border=1, fill=True, ln=0,align="C")
    document.cell(40, 10, "", border=1, fill=True, ln=1,align="C")
    document.set_text_color(0, 0, 0)
    for row in data:
        c_count = 0
        for item in row:
            document.cell(40, 5, txt=str(item), border=1, ln=0, align="C")
        document.cell(40,5,ln=1)
    document.cell(180)
    document.set_font('Arial', 'I', 10)
    perc = round((cand[i]["t_marks"]/cand[i]["median"])*100,2)
    if perc >= 100.0:
        perc = 100.00
    rem = round(100-perc,2)
    document.cell(40, 20, str(cand[i]["name"])+"'s overall percentile in the world is "+str(perc)+"%ile. This indicates that "+cand[i]["name"] + " has scored more than " + str(perc) + "% of students in the World and lesser than " + str(rem) + "% of students in the world", align="C", ln=1)
    document.cell(150)
    document.add_page(orientation="L")
    document.image('back.png', x=0, y=0, w=420, h=297)
    document.set_font('Arial', 'B', 15)
    document.rect(5.0, 5.0, 410.0, 287.0)
    document.rect(8.0, 8.0, 404.0, 282.0)
    document.cell(180)
    document.set_font('Arial', 'B', 15)
    document.cell(40, 20, 'Overview', align="C", ln=1)
    document.ln(20)
    document.set_font('Arial', '', 10)
    document.set_text_color(255, 255, 255)
    document.cell(20)
    document.cell(80, 10, "Average score of all students across the World", border=1, fill=True, ln=0, align="C")
    document.set_text_color(0, 0, 0)
    document.cell(20, 10,str(cand[i]["avg"]) , border=1, ln=0, align="C")
    document.cell(10)
    document.set_text_color(255, 255, 255)
    document.cell(100, 10, cand[i]["name"]+"'s attempts (Attempts x 100 / Total Questions)", border=1, fill=True, ln=0, align="C")
    document.set_text_color(0, 0, 0)
    document.cell(20, 10, str(cand[i]["attempt"]), border=1, ln=0, align="C")
    document.cell(10)
    document.set_text_color(255, 255, 255)
    document.cell(100, 10, cand[i]["name"] + "'s  Accuracy ( Corrects x 100 /Attempts )", border=1, fill=True,ln=0, align="C")
    document.set_text_color(0, 0, 0)
    document.cell(20, 10, str(cand[i]["accu"]), border=1, ln=1, align="C")
    #2nd row
    document.set_text_color(255, 255, 255)
    document.cell(20)
    document.cell(80, 10, "Median score of all students across the World", border=1, fill=True, ln=0, align="C")
    document.set_text_color(0, 0, 0)
    document.cell(20, 10, str(cand[i]["median"]), border=1, ln=0, align="C")
    document.cell(10)
    document.set_text_color(255, 255, 255)
    document.cell(100, 10, "Average attempts of all students across the World", border=1, fill=True,
                  ln=0, align="C")
    document.set_text_color(0, 0, 0)
    document.cell(20, 10, str(cand[i]["avg_attempt"]), border=1, ln=0, align="C")
    document.cell(10)
    document.set_text_color(255, 255, 255)
    document.cell(100, 10, "Average accuracy of all students across the World", border=1, fill=True, ln=0,
                  align="C")
    document.set_text_color(0, 0, 0)
    document.cell(20, 10, str(cand[i]["avg_accu"]), border=1, ln=1, align="C")
    #3rd row
    document.set_text_color(255, 255, 255)
    document.cell(20)
    document.cell(80, 10, "Mode score of all students across World", border=1, fill=True, ln=0, align="C")
    document.set_text_color(0, 0, 0)
    document.cell(20, 10, str(cand[i]["mode"]), border=1, ln=1, align="C")
    document.ln(20)
    #GRAPHS
    g1 = str(cand[i]["reg_no"]) + "_1.png"
    g2 = str(cand[i]["reg_no"]) + "_2.png"
    g3 = str(cand[i]["reg_no"]) + "_3.png"
    document.image(g1, x=None, y=None, w=130, h=100, type='', link='')
    document.cell(20)
    document.image(g2, x=140, y=100, w=130, h=100, type='', link='')
    document.cell(20)
    document.image(g3, x=270, y=100, w=130, h=100, type='', link='')
    document.output(file_name)
    # document = FPDF(orientation='L', unit='mm', format='A4')
    print("Report card of a student has been generated successfully.")
for i in range(2, max_row + 1):
    if i == 2:
        continue
    else:
        if sheet.cell(row=i, column=1).value not in cand.keys():
            cand[sheet.cell(row=i, column=1).value] = {}
            cand[sheet.cell(row=i, column=1).value]['name'] = sheet.cell(row=i, column=5).value
            cand[sheet.cell(row=i, column=1).value]['t_marks'] = int(sheet.cell(row=i, column=19).value)
            cand[sheet.cell(row=i, column=1).value]['reg_no'] = sheet.cell(row=i, column=6).value
            cand[sheet.cell(row=i, column=1).value]['grade'] = sheet.cell(row=i, column=7).value
            cand[sheet.cell(row=i, column=1).value]['gender'] = sheet.cell(row=i, column=9).value
            cand[sheet.cell(row=i, column=1).value]['dob'] = sheet.cell(row=i, column=10).value.date()
            cand[sheet.cell(row=i, column=1).value]['city'] = sheet.cell(row=i, column=11).value
            cand[sheet.cell(row=i, column=1).value]['dot'] = sheet.cell(row=i, column=12).value
            cand[sheet.cell(row=i, column=1).value]['country'] = sheet.cell(row=i, column=13).value
            cand[sheet.cell(row=i, column=1).value]['award'] = sheet.cell(row=i, column=20).value
            cand[sheet.cell(row=i, column=1).value]['questions'] =[["Question No","Attempt \nStatus",sheet.cell(row=i, column=5).value+"'s \nchoice","Correct \nAnswer","Outcome","Score if correct",sheet.cell(row=i, column=5).value+"'s score"]]
            cand[sheet.cell(row=i, column=1).value]['questions'].append([sheet.cell(row=i, column=14).value,"Unattempted" if sheet.cell(row=i, column=17).value == "Unattempted" else "Attempted",sheet.cell(row=i, column=15).value,sheet.cell(row=i, column=16).value,sheet.cell(row=i, column=17).value,sheet.cell(row=i, column=18).value,sheet.cell(row=i, column=19).value])
            cand[sheet.cell(row=i, column=1).value]['sec_2'] = [([sheet.cell(row=i, column=14).value,"Unattempted" if sheet.cell(row=i,column=17).value == "Unattempted" else "Attempted",sheet.cell(row=i, column=15).value,sheet.cell(row=i, column=16).value,sheet.cell(row=i, column=17).value,sheet.cell(row=i, column=19).value,sheet.cell(row=i, column=21).value,sheet.cell(row=i, column=22).value,sheet.cell(row=i, column=23).value,sheet.cell(row=i, column=24).value])]
            cand[sheet.cell(row=i, column=1).value]['school_name'] = sheet.cell(row=i, column=8).value
            cand[sheet.cell(row=i, column=1).value]['avg'] = sheet.cell(row=i, column=25).value
            cand[sheet.cell(row=i, column=1).value]['median'] = sheet.cell(row=i, column=26).value
            cand[sheet.cell(row=i, column=1).value]['mode'] = sheet.cell(row=i, column=27).value
            cand[sheet.cell(row=i, column=1).value]['attempt'] = sheet.cell(row=i, column=28).value
            cand[sheet.cell(row=i, column=1).value]['avg_attempt'] = sheet.cell(row=i, column=29).value
            cand[sheet.cell(row=i, column=1).value]['accu'] = round(100*sheet.cell(row=i, column=30).value,2)
            cand[sheet.cell(row=i, column=1).value]['avg_accu'] = round(100*sheet.cell(row=i, column=31).value,2)
        else:
            cand[sheet.cell(row=i, column=1).value]['t_marks'] = cand[sheet.cell(row=i, column=1).value]['t_marks'] + int(sheet.cell(row=i, column=19).value)
            cand[sheet.cell(row=i, column=1).value]['questions'].append([sheet.cell(row=i, column=14).value,
                                                                    "Unattempted" if sheet.cell(row=i, column=17).value == "Unattempted" else "Attempted",
                                                                    sheet.cell(row=i, column=15).value,
                                                                    sheet.cell(row=i, column=16).value,
                                                                    sheet.cell(row=i, column=17).value,
                                                                    sheet.cell(row=i, column=18).value,
                                                                    sheet.cell(row=i, column=19).value])
            cand[sheet.cell(row=i, column=1).value]['sec_2'].append([sheet.cell(row=i, column=14).value,"Unattempted" if sheet.cell(row=i,column=17).value == "Unattempted" else "Attempted",sheet.cell(row=i, column=15).value,sheet.cell(row=i, column=16).value,sheet.cell(row=i, column=17).value,sheet.cell(row=i, column=19).value,sheet.cell(row=i, column=21).value,sheet.cell(row=i, column=22).value,sheet.cell(row=i, column=23).value,sheet.cell(row=i, column=24).value])
count = 1
print("Welcome to Report Card Generation Tool")
for i in cand:
    print("Generating "+str(count)+"/"+str(len(cand))+" report card please wait...")
    bar_chart(cand[i])
    createPdf(i)
    count+=1
print("Finished generating all the report cards")

