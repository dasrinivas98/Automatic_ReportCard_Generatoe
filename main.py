import os
from openpyxl import load_workbook
from fpdf import FPDF
import matplotlib
import matplotlib.pyplot as plt
from os.path import join
filepath = "Dummy.xlsx"
wb = load_workbook(filepath,data_only=True)
sheet = wb.active
max_row = sheet.max_row
max_column = sheet.max_column
cand = {}
#create the chart inthe form of png to insert into pdf
def donut_chart(id):
    pie_name = str(cand[i]["reg_no"])+".png"
    print(pie_name)
    value = cand[i]["t_marks"]
    print(value)
    if value >= 80:
        color = [0.4392, 0.6784, 0.2784]
    if (value < 80) and (value >= 60):
        color = [0.9569, 0.6941, 0.5137]
    if value < 60:
        color = [1.0000, 0.4118, 0.4118]
    my_circle = plt.Circle((0, 0), 0.8, color=[1, 1, 1])
    if value > 0:
        values = [value, 100 - value]
        plt.pie(values,
                wedgeprops={'linewidth': 0, 'edgecolor': 'white'}, colors = ["tab:blue", "tab:orange"],labels = ["Correct answers","Wrong answers"], labeldistance=1.1,autopct='%.0f%%')
    p = plt.gcf()
    p.gca().add_artist(my_circle)
    plt.savefig(pie_name, orientation='portrait', transparent=True, bbox_inches=None, pad_inches=0)
    plt.close()
# create the pdf for the given data
def createPdf(id):
    pie_name = str(cand[i]["reg_no"]) + ".png"
    name = cand[i]["name"]
    ext = name + "_reportcard.pdf"
    desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
    report_cards = os.path.join(desktop,"report_cards")
    if not os.path.exists(report_cards):
        os.makedirs(report_cards)
    file_name = os.path.join(report_cards, ext)
    document = FPDF()
    document.add_page()
    document.set_font('Arial', 'B', 15)
    document.rect(5.0, 5.0, 200.0, 287.0)
    document.rect(8.0, 8.0, 194.0, 282.0)
    document.cell(80)
    document.cell(40, 10, 'PQR Championship Test', 0, 0, 'C')
    document.ln(10)
    document.cell(70)
    document.cell(60, 10, 'Report Card', 0, 0, 'C')
    document.ln(20)
    document.image("logo.png", 176, 10, w=20, h=20)
    document.image(name+".png",10,40,w=40,h=40)
    document.cell(50,50)
    document.cell(10,15,txt="Name : " + name, ln=0)
    document.cell(50)
    document.cell(10, 15, txt="Reg No. : " + str(cand[i]["reg_no"]), ln=1)
    document.cell(50)
    document.cell(10,15,txt="School : " + str(cand[i]["school_name"]), ln=0)
    document.cell(50)
    document.cell(10, 15, txt="Grade : " + str(cand[i]["grade"]), ln=1)
    document.cell(50)
    document.cell(10, 15, txt="Gender : " + str(cand[i]["gender"]), ln=0)
    document.cell(50)
    document.cell(10, 15, txt="Country : " + str(cand[i]["country"]), ln=1)
    document.cell(75)
    document.cell(10, 45, txt="Summary of Result", ln=1)
    document.cell(10)
    document.cell(10, 10, txt=name + " has successfully completed PQR championship test and", ln=1)
    document.cell(40)
    document.cell(10, 15, txt="scored " + str(cand[i]["t_marks"]) + " out of 100 marks and therefore", ln=1)
    document.cell(10)
    document.cell(10, 10, txt='" '+cand[i]["award"] +' "', ln=1)
    document.cell(30)
    document.image(pie_name, x=None, y=None, w=130, h=100, type='', link='')
    document.add_page(orientation="L")
    document.set_font('Arial', 'B', 15)
    document.rect(5.0, 5.0, 287.0, 200.0)
    document.rect(8.0, 8.0, 282.0, 194.0)
    document.cell(120)
    document.cell(40, 10, 'Score card', 0, 0, 'C')
    document.ln(15)
    data = cand[i]["questions"]
    col_width = document.w / 7
    row_height = document.font_size*1.1
    for row in data:
        document.cell(12)
        for item in row:
            document.cell(col_width, row_height * 1,
                     txt=str(item), border=1)
        document.ln(row_height * 1)
    document.output(file_name)
    document = FPDF(orientation='L', unit='mm', format='A4')
    print("pdf has been created successfully....")
for i in range(1, max_row + 1):
    if i == 1:
        continue
    else:
        if sheet.cell(row=i, column=1).value not in cand.keys():
            cand[sheet.cell(row=i, column=1).value] = {}
            cand[sheet.cell(row=i, column=1).value]['name'] = sheet.cell(row=i, column=5).value
            cand[sheet.cell(row=i, column=1).value]['t_marks'] = int(sheet.cell(row=i, column=19).value)
            cand[sheet.cell(row=i, column=1).value]['reg_no'] = sheet.cell(row=i, column=6).value
            cand[sheet.cell(row=i, column=1).value]['grade'] = sheet.cell(row=i, column=7).value
            cand[sheet.cell(row=i, column=1).value]['gender'] = sheet.cell(row=i, column=9).value
            cand[sheet.cell(row=i, column=1).value]['country'] = sheet.cell(row=i, column=13).value
            cand[sheet.cell(row=i, column=1).value]['award'] = sheet.cell(row=i, column=20).value
            cand[sheet.cell(row=i, column=1).value]['questions'] =[["Q.No","Your answer","Correct answer","Outcome","Score if correct","Your score"]]
            cand[sheet.cell(row=i, column=1).value]['questions'].append([sheet.cell(row=i, column=14).value,sheet.cell(row=i, column=15).value,sheet.cell(row=i, column=16).value,sheet.cell(row=i, column=17).value,sheet.cell(row=i, column=18).value,sheet.cell(row=i, column=19).value])
            cand[sheet.cell(row=i, column=1).value]['school_name'] = sheet.cell(row=i, column=8).value

        else:
            cand[sheet.cell(row=i, column=1).value]['t_marks'] = cand[sheet.cell(row=i, column=1).value]['t_marks'] + int(sheet.cell(row=i, column=19).value)
            cand[sheet.cell(row=i, column=1).value]['questions'].append([sheet.cell(row=i, column=14).value,
                                                                    sheet.cell(row=i, column=15).value,
                                                                    sheet.cell(row=i, column=16).value,
                                                                    sheet.cell(row=i, column=17).value,
                                                                    sheet.cell(row=i, column=18).value,
                                                                    sheet.cell(row=i, column=19).value])
count = 1
print("Welcome to Report Card Generation Tool")
for i in cand:
    print("Generating " + str(count) + "/" + str(len(cand)) + " report card please wait...")
    donut_chart(cand[i])
    createPdf(i)
    count += 1
print("Finished generating all the report cards")
